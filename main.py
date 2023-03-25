from openpyxl import *

inventory_file = load_workbook("inventory.xlsx")


product_list = inventory_file["Sheet1"]

products_amount_per_supplier= {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Add new column to Sheet1 with name Total_value
product_list.cell(1,5).value = "Total_value"


for product_row in range(2, product_list.max_row+1):
  supplier_name = product_list.cell(product_row, 4).value
  inventory = product_list.cell(product_row, 2).value
  price = product_list.cell(product_row, 3).value
  product_number = product_list.cell(product_row, 1).value
  
  #List each company with respective product count
  if supplier_name in products_amount_per_supplier:
    products_amount_per_supplier[supplier_name] += 1
  else:
    products_amount_per_supplier[supplier_name] = 1
    
  #List each company with respective total inventory value
  if  supplier_name in total_value_per_supplier: 
    total_value_per_supplier[supplier_name]  += inventory * price
  else:
    total_value_per_supplier[supplier_name]  = inventory * price
    
  #List products with inventory less than 10
  if inventory < 10 :
    products_under_10_inv[int(product_number)] = int(inventory)
  
  #Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet
  product_list.cell(product_row,5).value = inventory * price
  
  #Save the change into a new spreadsheet
  inventory_file.save("inventory_with_total_value.xlsx")
  
print(products_amount_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

